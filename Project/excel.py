from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from unidecode import unidecode


class Excel(Workbook):
	def __init__(self, innovation_productions):
		super(Excel, self).__init__()

		self.innovation_productions = innovation_productions

	def create_first_sheet(self):
		ws = self.active  # First sheet
		ws.title = "Tabela"
		ws['A1'].value = 'Professor'
		ws['B1'].value = 'Pontuação'
		for pos, researcher in enumerate(self.innovation_productions['Researcher']):
			ws[f'A{pos + 2}'].value = researcher

			df = self.innovation_productions['Productions'][pos]
			ws[f'B{pos + 2}'].value = f"='{researcher}'!E{df.shape[0]+2}"

	def create_sheets(self):
		for pos, researcher in enumerate(self.innovation_productions['Researcher']):
			ws = self.create_sheet(researcher)  # New sheet

			for row in dataframe_to_rows(self.innovation_productions['Productions'][pos], index=False, header=True):  # Add dataframe to the sheet
				ws.append(row)
		
	def add_score_formula(self):
		for ws in self.worksheets:
			if ws.title != 'Tabela':
				for pos, cell in enumerate(ws['E']):
					if pos != 0:
						cell.value = f'=C{pos + 1}*D{pos + 1}'

				last_row = len(ws['E'])
				ws[f"D{last_row + 1}"].value = "Total"
				ws[f"D{last_row + 1}"].font = Font(bold=True)  
				ws[f"E{last_row + 1}"].value = f"=SUM(E2:E{last_row})"
				ws[f"E{last_row + 1}"].font = Font(bold=True)  

	def format_as_table(self):
		for pos, ws in enumerate(self.worksheets):
			if ws.title != 'Tabela':
				dataframe = self.innovation_productions['Productions'][pos - 1]

				if not dataframe.empty:
					columns_dic = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E'}  # Get the alphabet's letter for the max number of columns
					last_column = columns_dic[len(dataframe.columns)]
					table_name = unidecode(self.innovation_productions['Researcher'][pos - 1].title().replace(' ', ''))
					table_ref = f"A1:{last_column}{dataframe.shape[0]+1}"  # The cells that must be converted to a table

					tab = Table(displayName=table_name, ref=table_ref)  # Create table based on cells reference
					style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)  # Table style
					
					tab.tableStyleInfo = style  # Apply the style

					ws.add_table(tab)  # Add table to sheet
			else:
				tab = Table(displayName='Tabela', ref=f"A1:B{len(ws['B'])}")  # Create table based on cells reference
				style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)  # Table style
				
				tab.tableStyleInfo = style  # Apply the style

				ws.add_table(tab)  # Add table to sheet
	def set_dimensions(self):
		for ws in self.worksheets:
			if ws.title != 'Tabela':
				ws.column_dimensions['A'].width = 40  # set width
				ws.column_dimensions['B'].width = 50
				ws.column_dimensions['C'].width = 15
				ws.column_dimensions['D'].width = 18
				ws.column_dimensions['E'].width = 15
			else:
				ws.column_dimensions['A'].width = 30  # set width
				ws.column_dimensions['B'].width = 20

	def set_alignment(self):
		for ws in self.worksheets:  # for each sheet
			for cell in ws[1]:
				cell.alignment = Alignment(horizontal='center', vertical='center')
			if ws.title != 'Tabela':	
				for col in ws.iter_cols(3, 5):   # iterate (columns 3->C and 5->E)
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')
			else:
				for cell in ws['B']:
					cell.alignment = Alignment(horizontal='center', vertical='center')

	def build_file(self):
		self.create_first_sheet()
		self.create_sheets()
		self.add_score_formula()
		self.format_as_table()
		self.set_dimensions()
		self.set_alignment()

			